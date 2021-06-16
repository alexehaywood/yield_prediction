#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Fri Feb  7 12:18:00 2020

@author: pcxah5
"""

import pandas as pd
import os
import numpy as np
from collections import defaultdict
import openpyxl
from sklearn.preprocessing import StandardScaler
from sklearn.metrics import r2_score, mean_squared_error
from sklearn.model_selection import KFold
import pickle

# import sys
# sys.path.insert(1, '/home/alexehaywood/Documents/yield_prediction/yield_prediction/')

from tools.data.rdkit_tools import caluclate_tanimoto
from tools.machine_learning import model_selection
from tools.machine_learning.kernels import kernel
from tools.utils.plotting import plotting


def save_fig_to_excel(excel_file, sheet_name, image_file, text=None):
    wb =  openpyxl.load_workbook(excel_file)
    
    ws1 = wb.create_sheet(sheet_name)
    
    img = openpyxl.drawing.image.Image(image_file)
    img.height=img.height*0.1
    img.width=img.width*0.1
    
    if text is None:
        img.anchor = 'A1'
        ws1.add_image(img)
        ws1.row_dimensions[1].height = img.height*0.1
        ws1.column_dimensions['A'].width = img.width*0.1
        
    else:
        ws1['A1'] = text
        img.anchor = 'B1'
        ws1.add_image(img)
        ws1.row_dimensions[1].height = img.height*0.1
        ws1.column_dimensions['B'].width = img.width*0.1
    
    wb.save(excel_file)


n_jobs=32
class machine_learning():
    
    def __init__(self, X=None, y=None, models=None, param_grid=None,
                 X_train=None, y_train=None, X_test=None, y_test=None):
        self.X = X
        self.y = y
        self.models = models
        self.param_grid = param_grid
        self.X_train = X_train
        self.y_train = y_train
        self.X_test = X_test
        self.y_test = y_test
        
    def split_descriptors_in_sample(self):
        self.X_train, self.X_test, self.y_train, self.y_test = \
            model_selection.split_data.random_split(self.X, self.y)
    
    def split_descriptors_out_of_sample(self, rxn_component, molecule_test_list):
        self.X_train = self.X[~self.X.index.get_level_values(
            rxn_component).isin(molecule_test_list)]
        self.y_train = self.y[~self.y.index.get_level_values(
            rxn_component).isin(molecule_test_list)]
        self.X_test = self.X[self.X.index.get_level_values(
            rxn_component).isin(molecule_test_list)]
        self.y_test = self.y[self.y.index.get_level_values(
            rxn_component).isin(molecule_test_list)]
    
    def preprocess_quantum_descriptors(self, X_train=None, X_test=None):
        if X_train is None:
            X_train = self.X_train
        if X_test is None:
            X_test = self.X_test
            
        scaler = StandardScaler()
        scaler.fit(X_train)
        
        self.X_train = pd.DataFrame(
            scaler.transform(X_train),
            index=X_train.index
            )
        if X_test:
            self.X_test = pd.DataFrame(
                scaler.transform(X_test), 
                index=X_test.index
                )
        
    def preprocess_graph_descriptors(self, X_train=None, X_test=None,
                                     kernel_params={}):
        if X_train is None:
            X_train = self.X_train
        if X_test is None:
            X_test = self.X_test
            
        # Create graph kernel matricies for the train and test set.
        graph_kernel = kernel('WeisfeilerLehman')
        
        k_train, k_test = graph_kernel.multiple_descriptor_types(
            X_train, X_test, n_jobs=n_jobs, **kernel_params)
        
        self.X_train = pd.DataFrame(k_train, index=X_train.index)
        if X_test is not None:
            self.X_test = pd.DataFrame(k_test, index=X_test.index)
        
    def preprocess_fingerprint_descriptors(self, X_train=None, X_test=None):
        if X_train is None:
            X_train = self.X_train
        if X_test is None:
            X_test = self.X_test
            
        # Calculate Tanimoto Scores.
        k_train = 1
        k_test = 1
        if X_test is not None:
            if X_train.isnull().values.any() or X_test.isnull().values.any():
                for i in X_train:
                    train = []
                    test = []
                    for fp1 in X_train[i]:
                        row = []
                        for fp2 in X_train[i]:
                            if pd.isnull(fp1) and pd.isnull(fp2):
                                row.append(2)
                            elif pd.isnull(fp1) or pd.isnull(fp2):
                                row.append(1)
                            else:
                                row.append(caluclate_tanimoto(fp1, fp2)+1)
                        train.append(row)
                    
                    for fp2 in X_test[i]:
                        row = []
                        for fp1 in X_train[i]:
                            if pd.isnull(fp1) and pd.isnull(fp2):
                                row.append(2)
                            elif pd.isnull(fp1) or pd.isnull(fp2):
                                row.append(1)
                            else:
                                row.append(caluclate_tanimoto(fp1, fp2)+1)
                        test.append(row)
                        
                    train = np.array(train)
                    test = np.array(test)
                    
                    k_train = k_train * train
                    k_test = k_test * test
                     
            else:
                for i in X_train:
                    train = []
                    test = []
                    for fp1 in X_train[i]:
                        row = []
                        for fp2 in X_train[i]:
                            row.append(caluclate_tanimoto(fp1, fp2))
                        train.append(row)
                    
                    for fp2 in X_test[i]:
                        row = []
                        for fp1 in X_train[i]:
                            row.append(caluclate_tanimoto(fp1, fp2))
                        test.append(row)
                        
                    train = np.array(train)
                    test = np.array(test)
            
                    k_train = k_train * train
                    k_test = k_test * test
        
        else:
            if X_train.isnull().values.any():
                for i in X_train:
                    train = []
                    for fp1 in X_train[i]:
                        row = []
                        for fp2 in X_train[i]:
                            if pd.isnull(fp1) and pd.isnull(fp2):
                                row.append(2)
                            elif pd.isnull(fp1) or pd.isnull(fp2):
                                row.append(1)
                            else:
                                row.append(caluclate_tanimoto(fp1, fp2)+1)
                        train.append(row)
                    
                    train = np.array(train)
                    
                    k_train = k_train * train
                
            else:
                for i in X_train:
                    train = []
                    for fp1 in X_train[i]:
                        row = []
                        for fp2 in X_train[i]:
                            row.append(caluclate_tanimoto(fp1, fp2))
                        train.append(row)
                    
                    train = np.array(train)
            
                    k_train = k_train * train

        self.X_train = pd.DataFrame(k_train, index=X_train.index)
        if X_test is not None:
            self.X_test = pd.DataFrame(k_test, index=X_test.index)
        
    def r2_and_rmse_scorer(self):
        rmse_scorer = model_selection.scorer(
            scorer_name_from_metrics_module='mean_squared_error'
            ).make(greater_is_better=False,squared=False)
        
        scoring = {'R-squared': 'r2',
                   'RMSE': rmse_scorer
                   }
        
        return scoring
        
    def tune_hyperparameters_and_run(self, 
            X_train=None, y_train=None, X_test=None, y_test=None, 
            save_plot_dir=None, save_table_dir=None, save_model_dir=None
            ):
        
        if self.param_grid is None:
            raise ValueError('"param_grid" is None. Please set "param_grid".')
            
        if X_train is None:    
            X_train = self.X_train
        if y_train is None:
            y_train = self.y_train
        if X_test is None:
            X_test = self.X_test
        if y_test is None:
            y_test = self.y_test
        
        tuned_models = defaultdict()
        best_params = defaultdict()
        best_cv_score = defaultdict()
        best_estimator = defaultdict()
        yield_pred = pd.DataFrame()
        scores = defaultdict()
        training_score = defaultdict(dict)
        
        if X_test is None:
            for model_name, model in self.models.items():
                print('Model: {} \n'.format(model_name))

                tuned_models[model_name], \
                best_params[model_name], \
                best_cv_score[model_name], \
                best_estimator[model_name], \
                    = model.tune_hyperparameters(
                        'GridSearchCV',
                        X_train.to_numpy(),
                        y_train,
                        X_test=X_test,
                        param_grid=self.param_grid[model_name],
                        scoring = self.r2_and_rmse_scorer(),
                        refit='R-squared',
                        n_jobs=n_jobs,
                        cv=KFold(n_splits=5, shuffle=True, random_state=0)
                        )
                
                best_cv_score[model_name]['mean_test_RMSE'] \
                    = best_cv_score[model_name]['mean_test_RMSE']*-1
                    
            # Save model.
            if save_model_dir is not None:
                for model_name, tuned_model in tuned_models.items():
                    pickle.dump(
                        tuned_model, 
                        open(
                            '{}/{}.joblib'.format(
                                save_model_dir, 
                                model_name.replace(' ', '_')
                                ), 
                             'wb'
                             )
                        )
            # Save results.
            if save_table_dir is not None:
                with pd.ExcelWriter( '{}/results.xlsx'.format(save_table_dir)) as writer:
    
                    for name, results in zip(
                            ['best_params', 'best_cv_score'],
                            [best_params, best_cv_score]
                            ):
    
                        results = pd.DataFrame(results).round(3)
                        results.to_excel(writer, sheet_name=name)
        
            return best_params, best_cv_score, best_estimator
            
        if y_test is None:
            for model_name, model in self.models.items():
                print('Model: {} \n'.format(model_name))

                tuned_models[model_name], \
                best_params[model_name], \
                best_cv_score[model_name], \
                best_estimator[model_name], \
                y_pred \
                    = model.tune_hyperparameters(
                        'GridSearchCV',
                        X_train.to_numpy(),
                        y_train,
                        X_test=X_test.to_numpy(),
                        param_grid=self.param_grid[model_name],
                        scoring = self.r2_and_rmse_scorer(),
                        refit='R-squared',
                        n_jobs=n_jobs
                        )
                y_pred = pd.DataFrame(
                    y_pred, index=X_test.index, 
                    columns=['yield_pred_{}'.format(model_name)]
                    )
                yield_pred = pd.concat([yield_pred, y_pred], axis=1)

                best_cv_score[model_name]['mean_test_RMSE'] \
                    = best_cv_score[model_name]['mean_test_RMSE']*-1
                    
                y_pred_train = model.tuned_model.predict(X_train.to_numpy())
                training_score[model_name]['R-squared'] = r2_score(y_train, y_pred_train)
                training_score[model_name]['RMSE'] = mean_squared_error(y_train, y_pred_train, squared=False)
            
            training_score = pd.DataFrame.from_dict(training_score)
            
            # Save model.
            if save_model_dir is not None:
                for model_name, tuned_model in tuned_models.items():
                    pickle.dump(
                        tuned_model, 
                        open(
                            '{}/{}.joblib'.format(
                                save_model_dir, 
                                model_name.replace(' ', '_')
                                ), 
                             'wb'
                             )
                        )
            # Save results.
            if save_table_dir is not None:
                with pd.ExcelWriter( '{}/results.xlsx'.format(save_table_dir)) as writer:
    
                    for name, results in zip(
                            ['best_params', 'best_cv_score', 'training_score'],
                            [best_params, best_cv_score, training_score]
                            ):
    
                        results = pd.DataFrame(results).round(3)
                        results.to_excel(writer, sheet_name=name)
    
                    yield_pred.to_excel(writer, sheet_name='y_pred', merge_cells=False)
    
            return best_params, best_cv_score, best_estimator, yield_pred
            
        elif y_test is not None:
            for model_name, model in self.models.items():
                print('Model: {} \n'.format(model_name))
                
                tuned_models[model_name], \
                best_params[model_name], \
                best_cv_score[model_name], \
                best_estimator[model_name], \
                y_pred, \
                scores[model_name] \
                    = model.tune_hyperparameters(
                        'GridSearchCV', 
                        X_train.to_numpy(), 
                        y_train, 
                        X_test=X_test.to_numpy(),
                        y_test=y_test,
                        param_grid=self.param_grid[model_name], 
                        scoring = self.r2_and_rmse_scorer(),
                        refit='R-squared',
                        n_jobs=n_jobs
                        )
                
                y_pred = pd.DataFrame(y_pred, index=X_test.index, 
                                      columns=['yield_pred_{}'.format(model_name)]
                                      )
                yield_pred = pd.concat([yield_pred, y_pred], axis=1)
                
                best_cv_score[model_name]['mean_test_RMSE'] \
                    = best_cv_score[model_name]['mean_test_RMSE']*-1
                scores[model_name]['RMSE'] = scores[model_name]['RMSE']*-1
                
                y_pred_train = model.tuned_model.predict(X_train.to_numpy())
                training_score[model_name]['R-squared'] = r2_score(y_train, y_pred_train)
                training_score[model_name]['RMSE'] = mean_squared_error(y_train, y_pred_train, squared=False)
            
            training_score = pd.DataFrame.from_dict(training_score)
                
            if save_plot_dir is not None:
                for model_name, model in self.models.items():
                    plotter = plotting(
                        rcParams={'font.size':10, 'axes.titlesize':10},
                        fig_kw={'figsize':(6, 4), 'ncols':1, 'nrows':1, 'dpi':600}
                        )
                    plotter.add_plot(
                        x=self.y_test,
                        y=yield_pred['yield_pred_{}'.format(model_name)],
                        kind='scatter',
                        plot_kw={'color':'grey', 'marker':'.', 's':2.5, 
                                 'alpha':0.75},
                        text={'x':0.95, 'y':0.05, 'fontsize':6, 'ha':'right',
                              's':'$R^2$: {:.2f}\nRMSE: {:.1f}'.format(
                                  scores[model_name]['R-squared'],
                                  scores[model_name]['RMSE']
                                  )
                              },
                        xlabel='Experimental Yield (%)',
                        ylabel='Predicted Yield (%)'
                        )
                    plotter.add_plot(
                        x=[0,100],
                        y=[0,100],
                        plot_kw={'linestyle':'dashed', 'color':'black',
                            #'alpha':0.95, 
                            'linewidth':0.75}
                        )
                    plotter.save_plot('{}/{}'.format(
                        save_plot_dir, 
                        model_name.replace(' ', '_')
                        ))
                            
            # Save model.
            if save_model_dir is not None:
                for model_name, tuned_model in tuned_models.items():
                    pickle.dump(
                        tuned_model, 
                        open(
                            '{}/{}.joblib'.format(
                                save_model_dir, 
                                model_name.replace(' ', '_')
                                ), 
                            'wb'
                            )
                        )
                    
            # Save results.
            if save_table_dir  is not None:
                with pd.ExcelWriter( '{}/results.xlsx'.format(save_table_dir)) as writer:
                    
                    for name, results in zip(
                            ['best_params', 'best_cv_score', 'scores', 'training_scores'], 
                            [best_params, best_cv_score, scores, training_score]
                            ):
                        
                        results = pd.DataFrame(results).round(3)
                        results.to_excel(writer, sheet_name=name)
                        
                    yield_pred.to_excel(writer, sheet_name='y_pred', merge_cells=False)
            
            return best_params, best_cv_score, best_estimator, yield_pred, scores

    # def run(self, X_train=None, y_train=None, X_test=None, y_test=None, 
    #         save_plot_dir=None, save_table_dir=None):

    #     if X_train is None:    
    #         X_train = self.X_train
    #     if y_train is None:
    #         y_train = self.y_train
    #     if X_test is None:
    #         X_test = self.X_test
    #     if y_test is None:
    #         y_test = self.y_test
        
        
    #     scoring = self.r2_and_rmse_scorer()
        
    #     model_params = defaultdict()
    #     yield_pred = pd.DataFrame()
    #     scores = defaultdict()
        
    #     for model_name, model in self.models.items():

    #         model_params[model_name] = model.get_params()

    #         fitted_model = model.fit(X_train, y_train)
            
    #         y_pred = fitted_model.predict(X_test)
    #         y_pred = pd.DataFrame(y_pred, index=X_test.index, 
    #                               columns=['yield_pred_{}'.format(model_name)]
    #                               )
    #         yield_pred = pd.concat([yield_pred, y_pred], axis=1)
            
    #         scores[model_name] = model_selection.scorer.run(
    #             scoring, fitted_model, X_test, y_test, y_pred
    #             )
    #         scores[model_name]['RMSE'] = scores[model_name]['RMSE']*-1
          
    #         if save_plot_dir is None:
    #                 plotting.plot_scatter(
    #                 self.y_test, 
    #                 yield_pred['yield_pred_{}'.format(model_name)], 
    #                 'Experimental Yield (%)', 
    #                 'Predicted Yield (%)',  
    #                 text='\n'.join(
    #                     '{}: {}'.format(k, round(v, 2)) 
    #                     for k, v 
    #                     in scores[model_name].items()
    #                     ),
    #                 )
    #         else:
    #             plotting.plot_scatter(
    #                 self.y_test, 
    #                 yield_pred['yield_pred_{}'.format(model_name)], 
    #                 'Experimental Yield (%)', 
    #                 'Predicted Yield (%)',  
    #                 text='\n'.join(
    #                     '{}: {}'.format(k, round(v, 2)) 
    #                     for k, v 
    #                     in scores[model_name].items()
    #                     ),
    #                 saveas='{}/{}'.format(save_plot_dir, model_name)
    #                 )
        
    #     yield_pred = pd.concat([y_test, yield_pred], axis=1)
        
    #     # Save out-of-sample results.
    #     if save_table_dir  is not None:
    #         with pd.ExcelWriter('{}/results.xlsx'.format(save_table_dir)) as writer:
                
    #             for name, results in zip(
    #                     ['model_params', 'scores'],
    #                     [model_params, scores]
    #                     ):
                    
    #                 results = pd.DataFrame(results).round(3)
    #                 results.to_excel(writer, sheet_name=name)
                
    #             yield_pred.sort_index(
    #                 level=['additive', 'aryl_halide', 'base', 'ligand']
    #                 ).to_excel(writer, sheet_name='y_pred', merge_cells=False)
        
    #     return yield_pred, scores
        
    def calculate_individual_scores(self, y_test, y_pred, molecule_test_list, 
                                    molecule_keys, rxn_component, 
                                    save_plot_dir, save_table_dir):
        
        # Calculate the R-squared and RMSE scores for the individual additives.
        individual_scores = defaultdict()
        
        for model_name in self.models.keys():
            if model_name not in os.listdir(save_plot_dir):
                    os.mkdir('{}/{}'.format(save_plot_dir, model_name))
            
            y_test_mol = defaultdict()
            y_pred_mol = defaultdict()
            scores_mol = defaultdict()
            for mol in molecule_test_list:
                mol_key = '{} {}'.format(
                    rxn_component.replace('_', ' ').title(),
                    ''.join(map(str, 
                                [k for k,v 
                                 in molecule_keys[rxn_component].items() 
                                 if v == mol
                                 ])
                            )
                    )
                
                y_test_mol[mol_key] = y_test[y_test.index.get_level_values(
                    rxn_component).isin([mol])]['yield_exp']
                
                y_pred_mol[mol_key] = y_pred[
                    y_pred.index.get_level_values(rxn_component).isin([mol])
                    ]['yield_pred_{}'.format(model_name)]
                
                # Calculate R-squared.
                r2 = model_selection.scorer(
                        scorer_name_from_metrics_module='r2_score'
                        ).score(
                            y_test_mol[mol_key],
                            y_pred_mol[mol_key]
                            )
                
                # Calculate RMSE.
                rmse = model_selection.scorer(
                        scorer_name_from_metrics_module='mean_squared_error'
                        ).score(
                            y_test_mol[mol_key], 
                            y_pred_mol[mol_key],
                            squared=False
                            )
                
                # Put scores in dictionary.
                scores = { (mol_key, model_name): {'R-Squared': r2,
                                              'RMSE': rmse}
                          }
                    
                scores_mol[mol_key] = \
                     '\n'.join('{}: {}'.format(k,round(v,2)) 
                               for item in scores.values() 
                               for k,v in item.items())
                
                # Plot graphs.
                plotting.plot_scatter(
                    x=y_test_mol[mol_key], 
                    y=y_pred_mol[mol_key], 
                    x_label='Experimental Yield (%)', 
                    y_label='Predicted Yield (%)',  
                    text=scores_mol[mol_key],
                    saveas='{}/{}/{}'.format(
                        save_plot_dir, model_name, mol_key),
                    title=mol_key
                    )
                
                individual_scores.update(scores)
            
            # Plot muli scatter.
            plotting.plot_scatter_subplots(
                nrows=plotting.plot_grid[len(y_test_mol)]['nrows'], 
                ncols=plotting.plot_grid[len(y_test_mol)]['ncols'], 
                x=list(y_test_mol.values()), 
                y=list(y_pred_mol.values()),
                title_subplots=list(y_test_mol.keys()), 
                text_subplots=list(scores_mol.values()), 
                xlim_subplots=(0, 100), 
                ylim_subplots=(-10, 110), 
                x_label='Observed Yield (%)', 
                y_label='Predicted Yield (%)',
                saveas='{}/{}/all_{}'.format(
                    save_plot_dir, model_name, rxn_component)
                )
            
            save_fig_to_excel(
                '{}/results.xlsx'.format(save_table_dir), 
                sheet_name='{}_figures'.format(model_name),
                image_file='{}/{}/all_{}.png'.format(
                    save_plot_dir, model_name, rxn_component)
                )
            
        # Save table.
        individual_scores = pd.DataFrame(individual_scores).T
        
        with pd.ExcelWriter(
                '{}/results.xlsx'.format(save_table_dir), 
                mode='a'
                ) as writer:
            
            individual_scores = individual_scores.round(3)
            
            individual_scores.to_excel(
                writer, sheet_name='{}_scores'.format(rxn_component),
                merge_cells=True)    
        
        return individual_scores

models = {
    'SVR - Linear Kernel': model_selection.model_selector(
            'svm', 'SVR', kernel='linear'),
    'SVR - Poly Kernel': model_selection.model_selector(
            'svm', 'SVR', kernel='poly'),
    'SVR - RBF Kernel': model_selection.model_selector(
            'svm', 'SVR', kernel='rbf'),
    'SVR - Sigmoid Kernel': model_selection.model_selector(
            'svm', 'SVR', kernel='sigmoid'),
    'SVR - Precomputed Kernel': model_selection.model_selector(
            'svm', 'SVR', kernel='precomputed'),
    'Linear Regression': model_selection.model_selector(
            'linear_model', 'LinearRegression'),
    'k-Nearest Neighbours': model_selection.model_selector(
            'neighbors', 'KNeighborsRegressor'),
    'Bayes Generalised Linear Model': model_selection.model_selector(
            'linear_model', 'BayesianRidge'),
    'Random Forest': model_selection.model_selector(
            'ensemble', 'RandomForestRegressor'),
    'Gradient Boosting': model_selection.model_selector(
            'ensemble', 'GradientBoostingRegressor'),
    'Decision Tree': model_selection.model_selector(
            'tree', 'DecisionTreeRegressor'),
    }

param_grid = {
    'SVR - Linear Kernel': {
        'C': [1, 10, 100, 1000],
        'epsilon': [1, 5, 10]
        },
    'SVR - Poly Kernel': {
        'C': [1, 10, 100, 1000],
        'epsilon': [1, 5, 10]
        },
    'SVR - RBF Kernel': {
        'C': [1, 10, 100, 1000],
        'epsilon': [1, 5, 10]
        },
    'SVR - Sigmoid Kernel': {
        'C': [1, 10, 100, 1000],
        'epsilon': [1, 5, 10]
        },
    'SVR - Precomputed Kernel': {
        'C': [1, 10, 100, 1000],
        'epsilon': [1, 5, 10]
        },
    'Linear Regression': {
        'fit_intercept': ['True', 'False']
        },
    'k-Nearest Neighbours': {
        'n_neighbors': [5, 10, 15, 20],
        'weights': ['uniform', 'distance']
        },
    'Bayes Generalised Linear Model': {
        'alpha_1': [1e-4, 1e-6, 1e-8],
        'alpha_2': [1e-4, 1e-6, 1e-8],
        'lambda_1': [1e-4, 1e-6, 1e-8],
        'lambda_2': [1e-4, 1e-6, 1e-8]
        },
    'Random Forest': {
        'n_estimators': [250, 500, 750, 1000]
        },
    'Gradient Boosting': {
        'n_estimators': [250, 500, 750, 1000],
        'learning_rate': [0.05, 0.1, 0.15, 0.2]
        },
    'Decision Tree': {
        }
    }


def in_sample(X, y, models, param_grid, X_type, saveas, save_plots=False, 
        save_table=False, save_model=False, kwargs={}):
    """
    Perform the in-sample test.
    
    """
    print('\n#### IN-SAMPLE TEST STARTED ####' + 
          '\nDescriptor Type: {}'.format(X_type) +
          '\n{}'.format(saveas)
          )
    
    if saveas is not None:
        if not os.path.exists(saveas):
            os.makedirs(saveas)
    
    # Set descriptors, observables, models and possible parameters for model 
    # tuning.
    in_sample_test = machine_learning(X, y, models, param_grid)
    
    print('\nSTEP 1: Splitting descriptors.')    
    # Split descriptors into train and test sets.
    in_sample_test.split_descriptors_in_sample()
    
    print('\nSTEP 2: Preprocessing descriptors.')
    if X_type == 'graphs':
        # Create graph kernels for the train and test descriptor sets.
        in_sample_test.preprocess_graph_descriptors(kernel_params=kwargs)
    elif X_type == 'quantum':
        # Scale the train and test descriptor sets.
        in_sample_test.preprocess_quantum_descriptors()
    elif X_type == 'quantum_scaled':
        pass
    elif X_type == 'fps':
        if X.iloc[0].dtypes == int:
            pass
        elif X.iloc[0].dtypes == object:
            in_sample_test.preprocess_fingerprint_descriptors()
    
    print('\nStep 3: Tuning hyperparameters and predicting yeild.')
    # Tune the models hyperparameters and predict reaction yield, calculate 
    # scores, plot graphs and save table.
    if save_plots:
        save_plot_dir='{}/plots'.format(saveas)
        if not os.path.exists(save_plot_dir):
            os.makedirs(save_plot_dir)
    else:
        save_plot_dir=None
    if save_table:
        save_table_dir=saveas
    else:
        save_table_dir=None
    if save_model:
        save_model_dir='{}/models'.format(saveas)
        if not os.path.exists(save_model_dir):
            os.makedirs(save_model_dir)
    else:
        save_model_dir=None
        
    in_sample_results = defaultdict()
    in_sample_results['best_params'], \
    in_sample_results['best_cv_score'], \
    in_sample_results['best_estimator'], \
    in_sample_results['y_pred'], \
    in_sample_results['scores'] \
        = in_sample_test.tune_hyperparameters_and_run(
            save_plot_dir=save_plot_dir,
            save_table_dir=save_table_dir,
            save_model_dir=save_model_dir
            )
        
    print('\n#### FINISHED ####\n\n')
    
    return in_sample_results


def out_of_sample(
        X, y, models, param_grid, X_type, molecule_test_list, molecule_keys, 
        rxn_component, saveas=None, save_plots=False, 
        save_table=False, save_model=False, kwargs={}
        ):
    """
    Perform the out-of-sample test.
    
    """
    print('\n#### OUT-OF-SAMPLE TEST STARTED ####' + 
          '\nDescriptor Type: {}'.format(X_type) +
          '\nReaction Component: {}'.format(rxn_component) +
          '\n{}'.format(saveas)
          )
    
    if saveas is not None:
        if not os.path.exists(saveas):
            os.makedirs(saveas)
        
    # Set descriptors, observables, models and possible parameters for model 
    # tuning.
    out_of_sample_test = machine_learning(X, y, models, param_grid)
    
    print('\nSTEP 1: Splitting descriptors.')
    # Split descriptors into train and test sets.
    out_of_sample_test.split_descriptors_out_of_sample(
        rxn_component, molecule_test_list)

    print('\nSTEP 2: Preprocessing descriptors.')
    if X_type == 'graphs':
        # Create graph kernels for the train and test descriptor sets.
        out_of_sample_test.preprocess_graph_descriptors(kernel_params=kwargs)
    elif X_type == 'quantum':
        # Scale the train and test descriptor sets.
        out_of_sample_test.preprocess_quantum_descriptors()
    elif X_type == 'fps':
        if X.iloc[0].dtypes == int:
            pass
        elif X.iloc[0].dtypes == object:
            out_of_sample_test.preprocess_fingerprint_descriptors()
    
    print('\nStep 3: Tuning hyperparameters and predicting yeild.')
    # Tune the models hyperparameters and predict reaction yield, calculate 
    # scores, plot graphs, save table and model.
    if save_plots:
        save_plot_dir='{}/plots'.format(saveas)
        if not os.path.exists(save_plot_dir):
            os.makedirs(save_plot_dir)
    else:
        save_plot_dir=None
    if save_table:
        save_table_dir=saveas
    else:
        save_table_dir=None
    if save_model:
        save_model_dir='{}/models'.format(saveas)
        if not os.path.exists(save_model_dir):
            os.makedirs(save_model_dir)
    else:
        save_model_dir=None
        
    out_of_sample_results = defaultdict()
    out_of_sample_results['best_params'], out_of_sample_results['best_cv_score'], \
    out_of_sample_results['best_estimator'], out_of_sample_results['y_pred'], \
    out_of_sample_results['scores'] \
        = out_of_sample_test.tune_hyperparameters_and_run(
            save_plot_dir=save_plot_dir,
            save_table_dir=save_table_dir,
            save_model_dir=save_model_dir
            )
    
    # if len(molecule_test_list) > 1:
    #     out_of_sample_results['individual_score'] \
    #         = out_of_sample_test.calculate_individual_scores(
    #             y_test=out_of_sample_results['y_pred'], 
    #             y_pred=out_of_sample_results['y_pred'], 
    #             molecule_test_list=molecule_test_list,
    #             molecule_keys=molecule_keys,
    #             rxn_component=rxn_component,
    #             save_plot_dir='{}/out_of_sample/{}{}'.format(
    #                 dir_plots, rxn_component, saveas_name),
    #             save_table_dir='{}/out_of_sample/{}{}'.format(
    #                 dir_tables, rxn_component, saveas_name)
    #             )
    
    print('\n#### FINISHED ####\n\n')
    
    return out_of_sample_results

def predict(X_train, y_train, X_test, models, param_grid, X_type, 
            saveas=None, save_table=False, save_model=False, kwargs={}
            ):
    """
    Perform the validation test.
    
    """
    print('\n#### VALIDATION TEST STARTED ####' + 
          '\nDescriptor Type: {}'.format(X_type) +
          '\n{}'.format(saveas)
          )
    
    if saveas is not None:
        if not os.path.exists(saveas):
            os.makedirs(saveas)
    
    # Set descriptors, observables, models and possible parameters for model 
    # tuning.
    validation_test = machine_learning(
        X_train=X_train, y_train=y_train, X_test=X_test, 
        models=models, param_grid=param_grid
        )
    
    print('\nSTEP 1: Preprocessing descriptors.')
    if X_type == 'graphs':
        # Create graph kernels for the train and test descriptor sets.
        validation_test.preprocess_graph_descriptors(kernel_params=kwargs)
    elif X_type == 'quantum':
        # Scale the train and test descriptor sets.
        validation_test.preprocess_quantum_descriptors()
    elif 'fp' in X_type:
        if X_train.iloc[0].dtypes == int:
           pass
        elif X_train.iloc[0].dtypes == object:
            validation_test.preprocess_fingerprint_descriptors()
    
    print('\nStep 2: Tuning hyperparameters and predicting yield.')
    # Tune the models hyperparameters and predict reaction yield, calculate 
    # scores, plot graphs, save table and model.
    if save_table:
        save_table_dir=saveas
    else:
        save_table_dir=None
    if save_model:
        save_model_dir='{}/models'.format(saveas)
        if not os.path.exists(save_model_dir):
            os.makedirs(save_model_dir)
    else:
        save_model_dir=None
    validation_results = defaultdict()
    if X_test:
        validation_results['best_params'], validation_results['best_cv_score'], \
        validation_results['best_estimator'], validation_results['y_pred'], \
            = validation_test.tune_hyperparameters_and_run(
                save_plot_dir=None,
                save_table_dir=save_table_dir,
                save_model_dir=save_model_dir
                )
    else:
        validation_results['best_params'], validation_results['best_cv_score'], \
        validation_results['best_estimator'], \
            = validation_test.tune_hyperparameters_and_run(
                save_plot_dir=None,
                save_table_dir=save_table_dir,
                save_model_dir=save_model_dir
                )
    
    print('\n#### FINISHED ####\n\n')
    
    return validation_results
